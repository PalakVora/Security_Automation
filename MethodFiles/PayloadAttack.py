import requests
import openpyxl
import re
import time
from openpyxl import Workbook
#import ReadExcel as MM

#API_Excel_Location = "D:/Programming/Application Security/coe-application-security/DataFiles/API.xlsx"
#API_Excel_Sheet_Name = "Fortify_Impact"
#API_Module_Name = "DarkWebPartnerSite"
#Payload_Excel_Location = "D:/Programming/Application Security/coe-application-security/DataFiles/Payloads.xlsx"
#Payload_Sheet_Name = "TEST"
AttackVector = "{Parameter}"



def perform_get_payload_attack(API_Excel_Location, API_Excel_Sheet_Name, API_Module_Name, Payload_Excel_Location, Payload_Sheet_Name):
    result = []
    print("Opening Excel")
    API_Excel_WorkBook_Location = openpyxl.load_workbook(API_Excel_Location)
    print("Opened Excel")
    for API_SheetName in API_Excel_WorkBook_Location.worksheets:
        if API_SheetName.title == API_Excel_Sheet_Name:
            API_ActiveWorkSheet = API_Excel_WorkBook_Location[API_SheetName.title]
            print("Title of Sheet = " + API_ActiveWorkSheet.title)
            API_RowLength = API_SheetName.max_row
            API_ColumnLength = API_SheetName.max_column
            for i in range(1, API_RowLength + 1):
                API_RowContents = API_SheetName.cell(row=i, column=1)
                print("Row " + str(API_RowContents.row) + " = " + str(API_RowContents.value))
                if ((API_RowContents.value) == API_Module_Name):
                    print("Module Found : " + str(API_RowContents.value) + " in Row - " + str(
                        API_RowContents.row) + " of Worksheet - " + API_ActiveWorkSheet.title)
                else:
                    continue

                for j in range(1, API_ColumnLength + 1):
                    API_ColumnContents = API_SheetName.cell(row=API_RowContents.row, column=j)
                    print(API_ColumnContents.value, end="" + "\n")

                    API_HTTPMethod = (API_SheetName["B" + str(API_RowContents.row)])
                    print("HTTP Method = " + str(API_HTTPMethod.value))

                    API_Request_Protocol = (API_SheetName["C" + str(API_RowContents.row)])
                    print("Protocol = " + str(API_Request_Protocol.value) + "\n")

                    API_Request_BaseURL = (API_SheetName["D" + str(API_RowContents.row)])
                    print("BaseURL = " + str(API_Request_BaseURL.value) + "\n")

                    API_Request_RelativeURL = (API_SheetName["E" + str(API_RowContents.row)])
                    print("RelativeURL = " + str(API_Request_RelativeURL.value) + "\n")

                    API_Request_Body = (API_SheetName["F" + str(API_RowContents.row)])
                    print("Request_Body = " + str(API_Request_Body.value) + "\n")

                    #API_Request_Header = (API_SheetName["G" + str(API_RowContents.row)])
                    #print("Header = " + API_Request_Header.value + "\n")
                    API_Request_Header = {'Content-Type': 'application/json; charset=UTF-8'}
                    print("Header = " + str(API_Request_Header) + "\n")

                    API_Request_Cookie = (API_SheetName["H" + str(API_RowContents.row)])
                    print("Cookie = " + str(API_Request_Cookie.value) + "\n")

                    API_CompleteURL = str(API_Request_Protocol.value) + "://" + str(API_Request_BaseURL.value) + str(API_Request_RelativeURL.value)
                    print("CompleteURL = " + API_CompleteURL)
                    break

                if(API_HTTPMethod.value == 'GET'):
                        print('Executing GET Method')

                        if (AttackVector in API_CompleteURL):
                            print("Parameter Observed in Request URL")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(
                                        Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableURL = str(API_CompleteURL.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable URL = " + VulnerableURL)
                                        response = requests.get((VulnerableURL), data=API_Request_Body, headers=API_Request_Header,cookies=API_Request_Cookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)

                        elif (AttackVector in API_Request_Body.value):
                            print("Parameter Observed in Request Body")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(
                                        Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableBody = str(API_Request_Body.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable Body = " + VulnerableBody)
                                        response = requests.get(str(API_CompleteURL), data=VulnerableBody, headers=API_Request_Header, cookies=API_Request_Cookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)

                        elif (AttackVector in API_Request_Header):
                            print("Parameter Observed in Request Header")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableHeader = str(API_Request_Header.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable Header = " + VulnerableHeader)
                                        response = requests.get(str(API_CompleteURL), data=API_Request_Body, headers=VulnerableHeader, cookies=API_Request_Cookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)

                        elif (AttackVector in API_Request_Cookie):
                            print("Parameter Observed in Request Cookie")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableCookie = str(API_Request_Cookie.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable Cookie = " + VulnerableCookie)
                                        response = requests.get(str(API_CompleteURL), data=API_Request_Body, headers=API_Request_Header, cookies=VulnerableCookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)

                else:
                    {"HTTP Method is not GET"}
                break
    return result


def perform_post_payload_attack(API_Excel_Location, API_Excel_Sheet_Name, API_Module_Name, Payload_Excel_Location, Payload_Sheet_Name):
    result = []
    print("Opening Excel")
    API_Excel_WorkBook_Location = openpyxl.load_workbook(API_Excel_Location)
    print("Opened Excel")
    for API_SheetName in API_Excel_WorkBook_Location.worksheets:
        if API_SheetName.title == API_Excel_Sheet_Name:
            API_ActiveWorkSheet = API_Excel_WorkBook_Location[API_SheetName.title]
            print("Title of Sheet = " + API_ActiveWorkSheet.title)
            API_RowLength = API_SheetName.max_row
            API_ColumnLength = API_SheetName.max_column
            for i in range(1, API_RowLength + 1):
                API_RowContents = API_SheetName.cell(row=i, column=1)
                print("Row " + str(API_RowContents.row) + " = " + str(API_RowContents.value))
                if ((API_RowContents.value) == API_Module_Name):
                    print("Module Found : " + str(API_RowContents.value) + " in Row - " + str(
                        API_RowContents.row) + " of Worksheet - " + API_ActiveWorkSheet.title)
                else:
                    continue

                for j in range(1, API_ColumnLength + 1):
                    API_ColumnContents = API_SheetName.cell(row=API_RowContents.row, column=j)
                    print(API_ColumnContents.value, end="" + "\n")

                    API_HTTPMethod = (API_SheetName["B" + str(API_RowContents.row)])
                    print("HTTP Method = " + str(API_HTTPMethod.value))

                    API_Request_Protocol = (API_SheetName["C" + str(API_RowContents.row)])
                    print("Protocol = " + str(API_Request_Protocol.value) + "\n")

                    API_Request_BaseURL = (API_SheetName["D" + str(API_RowContents.row)])
                    print("BaseURL = " + str(API_Request_BaseURL.value) + "\n")

                    API_Request_RelativeURL = (API_SheetName["E" + str(API_RowContents.row)])
                    print("RelativeURL = " + str(API_Request_RelativeURL.value) + "\n")

                    API_Request_Body = (API_SheetName["F" + str(API_RowContents.row)])
                    print("Request_Body = " + str(API_Request_Body.value) + "\n")

                    #API_Request_Header = (API_SheetName["G" + str(API_RowContents.row)])
                    #print("Header = " + API_Request_Header.value + "\n")
                    API_Request_Header = {'Content-Type': 'application/json; charset=UTF-8'}
                    print("Header = " + str(API_Request_Header) + "\n")

                    API_Request_Cookie = (API_SheetName["H" + str(API_RowContents.row)])
                    print("Cookie = " + str(API_Request_Cookie.value) + "\n")

                    API_CompleteURL = str(API_Request_Protocol.value) + "://" + str(API_Request_BaseURL.value) + str(
                        API_Request_RelativeURL.value)
                    print("CompleteURL = " + API_CompleteURL)

                    break

                if (API_HTTPMethod.value == 'POST'):
                    print('Executing POST Method')

                    if (AttackVector in API_CompleteURL):
                        print("Parameter Observed in Request URL")
                        Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                        for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                            if Payload_SheetName.title == Payload_Sheet_Name:
                                print("Worksheet Found = " + Payload_SheetName.title)
                                Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                Payload_RowLength = Payload_SheetName.max_row
                                Payload_ColumnLength = Payload_SheetName.max_column
                                print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                                for i in range(2, Payload_RowLength + 1):
                                    Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                    print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                    VulnerableURL = str(API_CompleteURL.value).replace(AttackVector,str(Payload_RowContents.value))
                                    print("\n" + "Vulnerable URL = " + VulnerableURL)
                                    response = requests.post(str(VulnerableURL), data=API_Request_Body, headers=API_Request_Header, cookies=API_Request_Cookie.value)
                                    StatusCode = str(response.status_code)
                                    print("StatusCode = " + str(StatusCode))
                                    print("Response Body = " + str(response.text))
                                    print("Response Header = " + str(response.headers))
                                    print("Response Cookie = " + str(response.cookies))
                                    result.append(StatusCode)
                                    print(result)
                                    time.sleep(10)

                    elif (AttackVector in API_Request_Body.value):
                        print("Parameter Observed in Request Body")
                        Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                        for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                            if Payload_SheetName.title == Payload_Sheet_Name:
                                print("Worksheet Found = " + Payload_SheetName.title)
                                Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                Payload_RowLength = Payload_SheetName.max_row
                                Payload_ColumnLength = Payload_SheetName.max_column
                                print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                                for i in range(2, Payload_RowLength + 1):
                                    Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                    print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                    VulnerableBody = str(API_Request_Body.value).replace(AttackVector, str(Payload_RowContents.value))
                                    print("\n" + "Vulnerable Body = " + VulnerableBody)
                                    response = requests.post(str(API_CompleteURL), data=VulnerableBody, headers=API_Request_Header, cookies=API_Request_Cookie.value)
                                    StatusCode = str(response.status_code)
                                    print("StatusCode = " + str(StatusCode))
                                    print("Response Body = " + str(response.text))
                                    print("Response Header = " + str(response.headers))
                                    print("Response Cookie = " + str(response.cookies))
                                    result.append(StatusCode)
                                    print(result)
                                    time.sleep(10)

                    elif (AttackVector in API_Request_Header):
                        print("Parameter Observed in Request Header")
                        Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                        for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                            if Payload_SheetName.title == Payload_Sheet_Name:
                                print("Worksheet Found = " + Payload_SheetName.title)
                                Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                Payload_RowLength = Payload_SheetName.max_row
                                Payload_ColumnLength = Payload_SheetName.max_column
                                print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                                for i in range(2, Payload_RowLength + 1):
                                    Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                    print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                    VulnerableHeader = str(API_Request_Header.value).replace(AttackVector, str(Payload_RowContents.value))
                                    print("\n" + "Vulnerable Header = " + VulnerableHeader)
                                    response = requests.post(str(API_CompleteURL), data=API_Request_Body, headers=VulnerableHeader, cookies=API_Request_Cookie.value)
                                    StatusCode = str(response.status_code)
                                    print("StatusCode = " + str(StatusCode))
                                    print("Response Body = " + str(response.text))
                                    print("Response Header = " + str(response.headers))
                                    print("Response Cookie = " + str(response.cookies))
                                    result.append(StatusCode)
                                    print(result)
                                    time.sleep(10)

                    elif (AttackVector in API_Request_Cookie):
                        print("Parameter Observed in Request Cookie")
                        Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                        for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                            if Payload_SheetName.title == Payload_Sheet_Name:
                                print("Worksheet Found = " + Payload_SheetName.title)
                                Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                Payload_RowLength = Payload_SheetName.max_row
                                Payload_ColumnLength = Payload_SheetName.max_column
                                print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(
                                    Payload_RowLength - 1))
                                for i in range(2, Payload_RowLength + 1):
                                    Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                    print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                    VulnerableCookie = str(API_Request_Cookie.value).replace(AttackVector, str(Payload_RowContents.value))
                                    print("\n" + "Vulnerable Cookie = " + VulnerableCookie)
                                    response = requests.post(str(API_CompleteURL), data=API_Request_Body, headers=API_Request_Header, cookies=VulnerableCookie.value)
                                    StatusCode = str(response.status_code)
                                    print("StatusCode = " + str(StatusCode))
                                    print("Response Body = " + str(response.text))
                                    print("Response Header = " + str(response.headers))
                                    print("Response Cookie = " + str(response.cookies))
                                    result.append(StatusCode)
                                    print(result)
                                    time.sleep(10)
                else:
                    {"HTTP Method is not POST"}
                break
    return result




def perform_put_payload_attack(API_Excel_Location, API_Excel_Sheet_Name, API_Module_Name, Payload_Excel_Location, Payload_Sheet_Name):
    result = []
    print("Opening Excel")
    API_Excel_WorkBook_Location = openpyxl.load_workbook(API_Excel_Location)
    print("Opened Excel")
    for API_SheetName in API_Excel_WorkBook_Location.worksheets:
        if API_SheetName.title == API_Excel_Sheet_Name:
            API_ActiveWorkSheet = API_Excel_WorkBook_Location[API_SheetName.title]
            print("Title of Sheet = " + API_ActiveWorkSheet.title)
            API_RowLength = API_SheetName.max_row
            API_ColumnLength = API_SheetName.max_column
            for i in range(1, API_RowLength + 1):
                API_RowContents = API_SheetName.cell(row=i, column=1)
                print("Row " + str(API_RowContents.row) + " = " + str(API_RowContents.value))
                if ((API_RowContents.value) == API_Module_Name):
                    print("Module Found : " + str(API_RowContents.value) + " in Row - " + str(
                        API_RowContents.row) + " of Worksheet - " + API_ActiveWorkSheet.title)
                else:
                    continue

                for j in range(1, API_ColumnLength + 1):
                    API_ColumnContents = API_SheetName.cell(row=API_RowContents.row, column=j)
                    print(API_ColumnContents.value, end="" + "\n")

                    API_HTTPMethod = (API_SheetName["B" + str(API_RowContents.row)])
                    print("HTTP Method = " + str(API_HTTPMethod.value))

                    API_Request_Protocol = (API_SheetName["C" + str(API_RowContents.row)])
                    print("Protocol = " + str(API_Request_Protocol.value) + "\n")

                    API_Request_BaseURL = (API_SheetName["D" + str(API_RowContents.row)])
                    print("BaseURL = " + str(API_Request_BaseURL.value) + "\n")

                    API_Request_RelativeURL = (API_SheetName["E" + str(API_RowContents.row)])
                    print("RelativeURL = " + str(API_Request_RelativeURL.value) + "\n")

                    API_Request_Body = (API_SheetName["F" + str(API_RowContents.row)])
                    print("Request_Body = " + str(API_Request_Body.value) + "\n")

                    #API_Request_Header = (API_SheetName["G" + str(API_RowContents.row)])
                    #print("Header = " + API_Request_Header.value + "\n")

                    API_Request_Header = {'Content-Type': 'application/json; charset=UTF-8'}
                    print("Header = " + str(API_Request_Header) + "\n")

                    API_Request_Cookie = (API_SheetName["H" + str(API_RowContents.row)])
                    print("Cookie = " + str(API_Request_Cookie.value) + "\n")

                    API_CompleteURL = str(API_Request_Protocol.value) + "://" + str(API_Request_BaseURL.value) + str(API_Request_RelativeURL.value)
                    print("CompleteURL = " + API_CompleteURL)

                    break

                if (AttackVector in API_CompleteURL):
                    print("Parameter Observed in Request URL")
                    Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                    for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                        if Payload_SheetName.title == Payload_Sheet_Name:
                            print("Worksheet Found = " + Payload_SheetName.title)
                            Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                            print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                            Payload_RowLength = Payload_SheetName.max_row
                            Payload_ColumnLength = Payload_SheetName.max_column
                            print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                            for i in range(2, Payload_RowLength + 1):
                                Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                VulnerableURL = str(API_CompleteURL.value).replace(AttackVector,str(Payload_RowContents.value))
                                print("\n" + "Vulnerable URL = " + VulnerableURL)
                                response = requests.put(str(VulnerableURL), data=API_Request_Body, headers=API_Request_Header, cookies=API_Request_Cookie.value)
                                StatusCode = str(response.status_code)
                                print("StatusCode = " + str(StatusCode))
                                print("Response Body = " + str(response.text))
                                print("Response Header = " + str(response.headers))
                                print("Response Cookie = " + str(response.cookies))
                                result.append(StatusCode)
                                print(result)
                                time.sleep(10)

                elif (AttackVector in API_Request_Body.value):
                    print("Parameter Observed in Request Body")
                    Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                    for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                        if Payload_SheetName.title == Payload_Sheet_Name:
                            print("Worksheet Found = " + Payload_SheetName.title)
                            Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                            print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                            Payload_RowLength = Payload_SheetName.max_row
                            Payload_ColumnLength = Payload_SheetName.max_column
                            print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                            for i in range(2, Payload_RowLength + 1):
                                Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                VulnerableBody = str(API_Request_Body.value).replace(AttackVector,str(Payload_RowContents.value))
                                print("\n" + "Vulnerable Body = " + VulnerableBody)
                                response = requests.put(str(API_CompleteURL), data=VulnerableBody, headers=API_Request_Header, cookies=API_Request_Cookie.value)
                                StatusCode = str(response.status_code)
                                print("StatusCode = " + str(StatusCode))
                                print("Response Body = " + str(response.text))
                                print("Response Header = " + str(response.headers))
                                print("Response Cookie = " + str(response.cookies))
                                result.append(StatusCode)
                                print(result)
                                time.sleep(10)

                elif (AttackVector in API_Request_Header):
                    print("Parameter Observed in Request Header")
                    Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                    for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                        if Payload_SheetName.title == Payload_Sheet_Name:
                            print("Worksheet Found = " + Payload_SheetName.title)
                            Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                            print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                            Payload_RowLength = Payload_SheetName.max_row
                            Payload_ColumnLength = Payload_SheetName.max_column
                            print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                            for i in range(2, Payload_RowLength + 1):
                                Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                VulnerableHeader = str(API_Request_Header.value).replace(AttackVector, str(Payload_RowContents.value))
                                print("\n" + "Vulnerable Header = " + VulnerableHeader)
                                response = requests.put(str(API_CompleteURL), data=API_Request_Body, headers=VulnerableHeader, cookies=API_Request_Cookie.value)
                                StatusCode = str(response.status_code)
                                print("StatusCode = " + str(StatusCode))
                                print("Response Body = " + str(response.text))
                                print("Response Header = " + str(response.headers))
                                print("Response Cookie = " + str(response.cookies))
                                result.append(StatusCode)
                                print(result)
                                time.sleep(10)

                elif (AttackVector in API_Request_Cookie):
                    print("Parameter Observed in Request Cookie")
                    Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                    for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                        if Payload_SheetName.title == Payload_Sheet_Name:
                            print("Worksheet Found = " + Payload_SheetName.title)
                            Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                            print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                            Payload_RowLength = Payload_SheetName.max_row
                            Payload_ColumnLength = Payload_SheetName.max_column
                            print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                            for i in range(2, Payload_RowLength + 1):
                                Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                VulnerableCookie = str(API_Request_Cookie.value).replace(AttackVector, str(Payload_RowContents.value))
                                print("\n" + "Vulnerable Cookie = " + VulnerableCookie)
                                response = requests.put(str(API_CompleteURL), data=API_Request_Body, headers=API_Request_Header, cookies=VulnerableCookie.value)
                                StatusCode = str(response.status_code)
                                print("StatusCode = " + str(StatusCode))
                                print("Response Body = " + str(response.text))
                                print("Response Header = " + str(response.headers))
                                print("Response Cookie = " + str(response.cookies))
                                result.append(StatusCode)
                                print(result)
                                time.sleep(10)
                else:
                    {"HTTP Method is not PUT"}
                break
    return result



def perform_allmethod_payload_attack(API_Excel_Location, API_Excel_Sheet_Name, API_Module_Name, Payload_Excel_Location, Payload_Sheet_Name):
    result = []

    print("Opening API Excel")
    API_Excel_WorkBook_Location = openpyxl.load_workbook(API_Excel_Location)
    print("Opened Excel")
    for API_SheetName in API_Excel_WorkBook_Location.worksheets:
        if API_SheetName.title == API_Excel_Sheet_Name:
            API_ActiveWorkSheet = API_Excel_WorkBook_Location[API_SheetName.title]
            print("Title of Sheet = " + API_ActiveWorkSheet.title)
            API_RowLength = API_SheetName.max_row
            API_ColumnLength = API_SheetName.max_column
            for i in range(1, API_RowLength + 1):
                API_RowContents = API_SheetName.cell(row=i, column=1)
                print("Row " + str(API_RowContents.row) + " = " + str(API_RowContents.value))
                if((API_RowContents.value) == API_Module_Name):
                    print("Module Found : " + str(API_RowContents.value) + " in Row - " + str(API_RowContents.row) + " of Worksheet - " + API_ActiveWorkSheet.title)
                else:
                    continue

                for j in range(1, API_ColumnLength + 1):
                    API_ColumnContents = API_SheetName.cell(row=API_RowContents.row, column=j)
                    print(API_ColumnContents.value, end="" + "\n")

                    API_HTTPMethod = (API_SheetName["B"+str(API_RowContents.row)])
                    print("HTTP Method = " + str(API_HTTPMethod.value))

                    API_Request_Protocol = (API_SheetName["C"+str(API_RowContents.row)])
                    print("Protocol = " + str(API_Request_Protocol.value) + "\n")

                    API_Request_BaseURL = (API_SheetName["D"+str(API_RowContents.row)])
                    print("BaseURL = " + str(API_Request_BaseURL.value) + "\n")

                    API_Request_RelativeURL = (API_SheetName["E"+str(API_RowContents.row)])
                    print("RelativeURL = " + str(API_Request_RelativeURL.value) + "\n")

                    API_Request_Body = (API_SheetName["F"+str(API_RowContents.row)])
                    print("Request_Body = " + str(API_Request_Body.value) + "\n")

                    #API_Request_Header = (API_SheetName["G"+str(API_RowContents.row)])
                    #print("Header = " + API_Request_Header.value + "\n")
                    API_Request_Header = {'Content-Type': 'application/json; charset=UTF-8'}
                    print("Header = " + str(API_Request_Header) + "\n")

                    API_Request_Cookie = (API_SheetName["H" + str(API_RowContents.row)])
                    print("Cookie = " + str(API_Request_Cookie.value) + "\n")

                    API_CompleteURL = str(API_Request_Protocol.value)+"://" + str(API_Request_BaseURL.value) + str(API_Request_RelativeURL.value)
                    print("CompleteURL = " + API_CompleteURL)

                    if(API_HTTPMethod.value=='POST'):
                        print('Executing POST Method')

                        if(AttackVector in API_CompleteURL):
                            print("Parameter Observed in Request URL")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableURL = str(API_CompleteURL.value).replace(AttackVector,str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable URL = " + VulnerableURL)
                                        response = requests.post(str(VulnerableURL), data=API_Request_Body, headers=API_Request_Header, cookies=API_Request_Cookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)

                        elif(AttackVector in API_Request_Body.value):
                            print("Parameter Observed in Request Body")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableBody = str(API_Request_Body.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable Body = " + VulnerableBody)
                                        response = requests.post(str(API_CompleteURL), data=VulnerableBody, headers=API_Request_Header, cookies=API_Request_Cookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)

                        elif(AttackVector in API_Request_Header):
                            print("Parameter Observed in Request Header")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableHeader = str(API_Request_Header.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable Header = " + VulnerableHeader)
                                        response = requests.post(str(API_CompleteURL), data=API_Request_Body, headers=VulnerableHeader, cookies=API_Request_Cookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)

                        elif(AttackVector in API_Request_Cookie):
                            print("Parameter Observed in Request Cookie")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(
                                        Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableCookie = str(API_Request_Cookie.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable Cookie = " + VulnerableCookie)
                                        response = requests.post(str(API_CompleteURL), data=API_Request_Body, headers=API_Request_Header, cookies=VulnerableCookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)
                        elif():
                            print("No Parameter Observed in " + str(API_RowContents.value))

                    elif (API_HTTPMethod.value == 'PUT'):
                        print('Executing POST Method')

                        if (AttackVector in API_CompleteURL):
                            print("Parameter Observed in Request URL")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(
                                        Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableURL = str(API_CompleteURL.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable URL = " + VulnerableURL)
                                        response = requests.put(str(VulnerableURL), data=API_Request_Body, headers=API_Request_Header,cookies=API_Request_Cookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)

                        elif (AttackVector in API_Request_Body.value):
                            print("Parameter Observed in Request Body")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(
                                        Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableBody = str(API_Request_Body.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable Body = " + VulnerableBody)
                                        response = requests.put(str(API_CompleteURL), data=VulnerableBody, headers=API_Request_Header, cookies=API_Request_Cookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)

                        elif (AttackVector in API_Request_Header):
                            print("Parameter Observed in Request Header")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableHeader = str(API_Request_Header.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable Header = " + VulnerableHeader)
                                        response = requests.put(str(API_CompleteURL), data=API_Request_Body, headers=VulnerableHeader, cookies=API_Request_Cookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)

                        elif (AttackVector in API_Request_Cookie):
                            print("Parameter Observed in Request Cookie")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableCookie = str(API_Request_Cookie.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable Cookie = " + VulnerableCookie)
                                        response = requests.put(str(API_CompleteURL), data=API_Request_Body, headers=API_Request_Header, cookies=VulnerableCookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)
                        elif ():
                            print("No Parameter Observed in " + str(API_RowContents.value))

                    elif (API_HTTPMethod.value == 'GET'):
                        print('Executing GET Method')

                        if (AttackVector in API_CompleteURL):
                            print("Parameter Observed in Request URL")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(
                                        Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableURL = str(API_CompleteURL.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable URL = " + VulnerableURL)
                                        response = requests.get((VulnerableURL), data=API_Request_Body, headers=API_Request_Header,cookies=API_Request_Cookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)

                        elif (AttackVector in API_Request_Body.value):
                            print("Parameter Observed in Request Body")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(
                                        Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableBody = str(API_Request_Body.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable Body = " + VulnerableBody)
                                        response = requests.get(str(API_CompleteURL), data=VulnerableBody, headers=API_Request_Header, cookies=API_Request_Cookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)

                        elif (AttackVector in API_Request_Header):
                            print("Parameter Observed in Request Header")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableHeader = str(API_Request_Header.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable Header = " + VulnerableHeader)
                                        response = requests.get(str(API_CompleteURL), data=API_Request_Body, headers=VulnerableHeader, cookies=API_Request_Cookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)

                        elif (AttackVector in API_Request_Cookie):
                            print("Parameter Observed in Request Cookie")
                            Payload_Excel_WorkBook_Location = openpyxl.load_workbook(Payload_Excel_Location)
                            for Payload_SheetName in Payload_Excel_WorkBook_Location.worksheets:
                                if Payload_SheetName.title == Payload_Sheet_Name:
                                    print("Worksheet Found = " + Payload_SheetName.title)
                                    Payload_ActiveWorkSheet = Payload_Excel_WorkBook_Location[Payload_SheetName.title]
                                    print("Title of Sheet = " + Payload_ActiveWorkSheet.title)
                                    Payload_RowLength = Payload_SheetName.max_row
                                    Payload_ColumnLength = Payload_SheetName.max_column
                                    print("Number of Payloads for " + str(Payload_SheetName.title) + " = " + str(Payload_RowLength - 1))
                                    for i in range(2, Payload_RowLength + 1):
                                        Payload_RowContents = Payload_SheetName.cell(row=i, column=1)
                                        print("Row " + str(Payload_RowContents.row - 1) + " = " + str(Payload_RowContents.value), end="" + "\n")
                                        VulnerableCookie = str(API_Request_Cookie.value).replace(AttackVector, str(Payload_RowContents.value))
                                        print("\n" + "Vulnerable Cookie = " + VulnerableCookie)
                                        response = requests.get(str(API_CompleteURL), data=API_Request_Body, headers=API_Request_Header, cookies=VulnerableCookie.value)
                                        StatusCode = str(response.status_code)
                                        print("StatusCode = " + str(StatusCode))
                                        print("Response Body = " + str(response.text))
                                        print("Response Header = " + str(response.headers))
                                        print("Response Cookie = " + str(response.cookies))
                                        result.append(StatusCode)
                                        print(result)
                                        time.sleep(10)
                        elif ():
                            print("No Parameter Observed in " + str(API_RowContents.value))
                    else:
                        print("Unable to execute Method")

                    break
                break
    return result


