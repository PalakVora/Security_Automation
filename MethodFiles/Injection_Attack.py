import openpyxl,json, traceback, logging
import sys
# insert at 1, 0 is the script path (or '' in REPL)
sys.path.insert(1, 'PythonFiles/Utilities')

from Get_AttackAPI import setAPI
from Request_File import requestQuery
from Check_condition import injectionCondition

keyword_payload = ["union","group"]

# ****************************** SQL INJECTION CONDITIONS ************************************
def condition_check(category,elapsed_time,after_hit_time,object1,after_hit,body_text,result,after_hit_result):
    
    object1.after_hit=after_hit
    object1.check_body=body_text
    object1.original_response=result
    object1.after_hit_res=after_hit_result
    if category == "Time":
        object1.category = "Time"
    else:
        object1.category=""
    object1.original_response_time=elapsed_time
    object1.response_time = after_hit_time
    object1.check_for_errors()
    return object1.flag


# ****************************** HIT API AND GET RESPONSE WITH PAYLOAD ************************************  
           

def hit_request(method,new_url,body,header,cookie):
    try:
        Object1=requestQuery()
        if method == "GET":
            print(new_url,body,header, cookie)
            hitreq , res ,time_elapsed = Object1.hit_get( new_url, body, header, cookie)
            return hitreq, res ,time_elapsed
        elif method == "POST":
            hitreq, res, time_elapsed = Object1.hit_post( new_url, body, header, cookie)
            return hitreq, res, time_elapsed 
        elif method == "PUT":
            hitreq, res,time_elapsed = Object1.hit_put( new_url, body, header, cookie)
            return hitreq, res,time_elapsed
        elif method == "DELETE":
            hitreq, res, time_elapsed = Object1.hit_delete( new_url, body, header, cookie)
            return hitreq, res, time_elapsed

    except:
        print("Failed to hit request")


# ****************************** INDEX OF TIME AND BLIND ************************************     
def find_index(payload_excel_workbook_location,attack_payload_sheetname):
    time_index=0
    blind_index=0
    for payload_sheetname in payload_excel_workbook_location.worksheets:
        if payload_sheetname.title == attack_payload_sheetname:
            for i in range(2, payload_sheetname.max_row):
                payload_rowcontents = payload_sheetname.cell(row=i,column=1)
                if str(payload_rowcontents.value) == "SQL Time":
                    time_index = i
                if str(payload_rowcontents.value) == "SQL Blind":
                    blind_index = i
        
    return time_index , blind_index  
# ************************************ form payload request **********************
def form_request(loaded_payload_replace,key,Url,Body,Header,Cookie):
    if key =='U':
        Url=loaded_payload_replace
    elif key == 'B':
        Body =loaded_payload_replace
    return Url,Body,Header,Cookie

# ********************************************* LOAD PAYLOAD ***********************************************
def load_payload(key,category,elapsed_time,object1,Method,replace_this,Url,Body,Header,Cookie,result,payload_rowcontents,attack_area):   
    #================================== Replace with payload ========================
    loaded_payload_replace = str(attack_area).replace(replace_this, str(payload_rowcontents.value))
    loaded_payload_replace = loaded_payload_replace.replace("$","")
    Url,Body,Header,Cookie=form_request(loaded_payload_replace,key,Url,Body,Header,Cookie)
    after_hit, after_hit_result,after_hit_time = hit_request(Method,Url,Body,Header,Cookie)
    flag = condition_check(category,elapsed_time,after_hit_time,object1,after_hit,after_hit.text,result,after_hit_result)
    #=================================== Append with union/group ============================
    keep_variable = str(payload_rowcontents.value)
    keep_variable=keep_variable.lower()           
    for x in keyword_payload:
        if(x in keep_variable): 
            replacing_with = replace_this + ' ' + str(payload_rowcontents.value)
            loaded_payload_append_space = str(attack_area).replace(replace_this, replacing_with)
            loaded_payload_append_space = loaded_payload_append_space.replace("$","")
            Url,Body,Header,Cookie=form_request(loaded_payload_append_space,key,Url,Body,Header,Cookie)
            after_hit, after_hit_result,after_hit_time = hit_request(Method,Url,Body,Header,Cookie)
            flag =condition_check(category,elapsed_time,after_hit_time,object1,after_hit,after_hit.text,result,after_hit_result)
            break
    #================================== Append ============================
    replacing_with = replace_this  + str(payload_rowcontents.value)
    loaded_payload_append = str(attack_area).replace(replace_this, replacing_with)
    loaded_payload_append = loaded_payload_append.replace("$","")
    Url,Body,Header,Cookie=form_request(loaded_payload_append,key,Url,Body,Header,Cookie)
    after_hit, after_hit_result,after_hit_time = hit_request(Method,Url,Body,Header,Cookie)
    flag =condition_check(category,elapsed_time,after_hit_time,object1,after_hit,after_hit.text,result,after_hit_result)
    #=================================== Append with space ============================
    replacing_with = replace_this + ' ' + str(payload_rowcontents.value)
    loaded_payload_space = str(attack_area).replace(replace_this, replacing_with)
    loaded_payload_space = loaded_payload_space.replace("$","")
    Url,Body,Header,Cookie=form_request(loaded_payload_space,key,Url,Body,Header,Cookie)
    after_hit, after_hit_result,after_hit_time = hit_request(Method,Url,Body,Header,Cookie)
    flag =condition_check(category,elapsed_time,after_hit_time,object1,after_hit,after_hit.text,result,after_hit_result)
    return flag    

# ***************************************** PAYLOAD SHEET , CALL FOR ATTACK **************************************            
def attack_it(key,elapsed_time,Method,attack_position,attack_area,Url,Body,Header,Cookie,result,payload_excel_location,attack_payload_sheetname):
    payload_excel_workbook_location=openpyxl.load_workbook(payload_excel_location)
    replace_this = ""
    flagA = 0
    flagB = 0
    flagC = 0
    flag = 0
    object1=injectionCondition()
    time_index ,blind_index = find_index(payload_excel_workbook_location,attack_payload_sheetname)
    replace_this = attack_position
    loaded_payload_replace = str(attack_area).replace(replace_this, "'")
    loaded_payload_replace = loaded_payload_replace.replace("$","")
    Url,Body,Header,Cookie=form_request(loaded_payload_replace,key,Url,Body,Header,Cookie)
    after_hit, after_hit_result,after_hit_time = hit_request(Method,Url,Body,Header,Cookie)
    flagC = condition_check("",elapsed_time,after_hit_time,object1,after_hit,after_hit.text,result,after_hit_result)
    for payload_sheetname in payload_excel_workbook_location.worksheets:
        if payload_sheetname.title == attack_payload_sheetname:
            payload_activeworksheet = payload_excel_workbook_location[payload_sheetname.title]
            payload_rowlength = payload_sheetname.max_row
            # ------------------------ SQL Error based ---------------------------------------
            for i in range(2, time_index):
                payload_rowcontents = payload_sheetname.cell(row=i, column=1)
                category = "Error"
                flagA = load_payload(key,category,elapsed_time,object1,Method,attack_position,Url, Body,Header,Cookie,result,payload_rowcontents,attack_area)
            # ------------------------- SQL Time Based ----------------------------------------------
            for j in range (time_index+1 , blind_index):
                payload_rowcontents = payload_sheetname.cell(row=i, column=1)
                category = "Time"
                flagB = load_payload(key,category,elapsed_time,object1,Method,attack_position,Url,Body,Header,Cookie,result,payload_rowcontents,attack_area)
           # for k in range (blind_index+1 ,payload_rowlength+1)
            #    payload_rowcontents = payload_sheetname.cell(row=i, column=1)
             #   print("Row " + i + " = " + str(payload_rowcontents.value), end="" + "\n")
              #  category = "Blind"
               # load_payload(category,elapsed_time,object1,Method,replace_this,Body,Header,Cookie,result,payload_rowcontents,attack_area)
    flag = flagA or flagB or flagC
    return flag        


# ****************************** IDENTIFY ATTACK AREA, HIT CORRECT REQUEST, CALL FOR SQL INJECTION ************************************

def API_injection(api_name,http_method,protocol,base_url,relative_url,request_body,header,cookies,payload_excel_location,attack_payload_sheetname):
    result = {}
    payload_param = []
    flag = 0
    elapsed_time = ""
    key=''
    position=0
    object1=setAPI()
    try:
        # Get API into variables
        
        
        object1.api_name=api_name
        object1.raw_http_method=http_method
        object1.raw_protocol=protocol
        object1.base_url=base_url
        object1.relative_url=relative_url
        object1.raw_request_body=json.loads(request_body)
        object1.raw_header=json.loads(header)
        object1.raw_cookies=json.loads(cookies)

        # Make the URL
        object1.make_url()

        #Replace dollar signs
        dollar_pos,payload_param=object1.replace_dollar()
    except Exception as error:
        print("Failed initial API assemble")
        
    
    # Hit original request and send for payload
    try:
        Object1=requestQuery()
        if(http_method == 'GET'):
            #Original request
            try:
                print(object1.url, object1.request_body, object1.header, object1.cookies)
                hitreq,res,elapsed_time=Object1.hit_get(object1.url, object1.request_body, object1.header, object1.cookies)
            except:
                print("Failed to hit request")
            
            #Identifying areas for attack and sending for attack
            for i in payload_param:
                if i == "URL":
                    key='U'
                    flag = attack_it(key,elapsed_time,http_method,dollar_pos[position],object1.raw_url,object1.url,object1.request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                    position=position+1
                if i == "Body":
                    key='B'
                    flag = attack_it(key,elapsed_time,http_method,dollar_pos[position],object1.raw_request_body,object1.url,object1.request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                    position=position+1
                if i == "Header":
                    flag =  attack_it(elapsed_time,http_method,dollar_pos[position],object1.raw_header,object1.url,object1.request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                    position=position+1
                if i == "Cookie":
                    flag =  attack_it(elapsed_time,http_method,dollar_pos[position],object1.raw_cookies,object1.url,object1.request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                else:
                    print("No position selected")

                
                        
                    
        elif (http_method == 'POST'):
            #==== Original request ==============
            abc,res,elapsed_time=Object1.hit_post(object1.url, object1.request_body, object1.header, object1.cookies)
            #============= Payload request==========
            for i in payload_param:
                if i == "URL":
                    key='U'
                    attack_it(key,elapsed_time,http_method,dollar_pos[position],object1.raw_url,object1.url,object1.request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                    position=position+1
                if i == "Body":
                    key='B'
                    attack_it(key,elapsed_time,http_method,dollar_pos[position],object1.raw_request_body,object1.url,object1.request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                    position=position+1
                if i == "Header":
                    attack_it(key,elapsed_time,http_method,dollar_pos[position],object1.raw_header,object1.url,object1.request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                    position=position+1
                if i == "Cookie":
                    attack_it(key,elapsed_time,http_method,dollar_pos[position],object1.raw_cookies,object1.url,object1.request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                else:
                    print("No position selected")

        elif(http_method == 'PUT'):
            abc, res,elapsed_time = Object1.hit_put(object1.url, object1.reques_body, object1.header, object1.cookies)
            for i in payload_param:
                if i == "URL":
                    print("inside for")
                    attack_it(key,elapsed_time,elapsed_time,http_method,dollar_pos,object1.raw_url,object1.url,object1.request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                if i == "Body":
                    attack_it(key,elapsed_time,elapsed_time,http_method,dollar_pos,object1.url,object1.raw_request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                if i == "Header":
                    attack_it(key,elapsed_time,elapsed_time,http_method,dollar_pos,object1.raw_header,object1.url,object1.request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                if i == "Cookie":
                    attack_it(key,elapsed_time,elapsed_time,http_method,dollar_pos,object1.raw_cookies,object1.url,object1.request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                else:
                    print("No position selected")

        elif(http_method == 'DELETE'):
            abc, result , elapsed_time= Object1.hit_delete(object1.url, object1.reques_body, object1.header, object1.cookies)
            for i in payload_param:
                if i == "URL":
                    attack_it(key,elapsed_time,elapsed_time,http_method,dollar_pos,object1.raw_url,object1.url,object1.request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                if i == "Body":
                    attack_it(key,elapsed_time,elapsed_time,http_method,dollar_pos,object1.url,object1.raw_request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                if i == "Header":
                    attack_it(key,elapsed_time,elapsed_time,http_method,dollar_pos,object1.raw_header,object1.url,object1.request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                if i == "Cookie":
                    attack_it(key,elapsed_time,elapsed_time,http_method,dollar_pos,object1.raw_cookies,object1.url,object1.request_body,object1.header,object1.cookies,res,payload_excel_location,attack_payload_sheetname)
                else:
                    print("No position selected")
    except Exception as error:
        print(error)
        traceback.print_stack()
    return flag


"""
if __name__ == "__main__":
    ole = API_injection(api_name,http_method,protocol,base_url,relative_url,request_body,header,cookies,payload_excel_location,attack_payload_sheetname)
api_name = "TestAPI"
http_method="POST"
protocol="https"
base_url="reqres.in"
relative_url="/api/$users$"
request_body='{"name": "$morpheus$", "job": "leader2"}'
header='{"Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8","Connection": "keep-alive","Referer": "https://reqres.in/","Upgrade-Insecure-Requests": "1"}'
cookies='{"":""}'
payload_excel_location = "D:/Programming/Application Security/coe-application-security/DataFiles/Payloads.xlsx"
attack_payload_sheetname = "Sheet3"
api_name = "TestAPIDefend"
http_method="GET"
protocol="https"
base_url="defendtheweb.net"
relative_url="/playground/sqli2?q=$A$"
request_body='{"":""}'
header='{"Content-Type": "application/json","Connection": "keep-alive","Cookie": "PHPSESSID=fmm8pac5gr2d68cb9dltd2lrql; cookies_dismissed=1"}'
cookies='{"":""}'
payload_excel_location = "D:/Programming/Application Security/coe-application-security/DataFiles/Payloads.xlsx"
attack_payload_sheetname = "SQL"
api_name = "TestAPI"
http_method="GET"
protocol="https"
base_url="reqres.in"
relative_url="/api/users/$2$"
request_body='{"":""}'
header='{"Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8","Connection": "keep-alive","Referer": "https://reqres.in/","Upgrade-Insecure-Requests": "1"}'
cookies='{"":""}'
payload_excel_location = "D:/Programming/Application Security/coe-application-security/DataFiles/Payloads.xlsx"
attack_payload_sheetname = "SQL"

if __name__ == "__main__":
    ole = API_wert(api_name,http_method,protocol,base_url,relative_url,request_body,header,cookies,payload_excel_location,attack_payload_sheetname)


if __name__ == "__main__":
    ole = API_wert(api_name,http_method,protocol,base_url,relative_url,request_body,header,cookies,payload_excel_location,attack_payload_sheetname)
if __name__ == "__main__":
    API_wert(Excel_Location, Excel_Sheet_Name, Module_Name,payload_excel_location,attack_payload_sheetname)

    try:
        print("Executing PDF")
        pdf = FPDF()
        print("Adding Page")
        pdf.add_page()
        pdf.set_margins(5, 5 , 5)
        print("Setting Font")
        pdf.set_font("Cambria", size=5)
        print("Setting Page Title")
        pdf.cell(200, 10, txt="COE SECURITY TEST REPORT", ln=2, align='C')


        print("Setting Second Line")
        pdf.cell(100, 5, txt='Status Code: ' + str(URL), ln=2, align='A')
        pdf.cell(100, 5, txt='Status Code: ' + result['StatusCode'], ln=1, align='A')
        print("Converting as Output")
        pdf.output("COE - Security Scan Report - " + API+".pdf")

    except Exception as error:
        print(error)
        traceback.print_stack()




        def readexcel(Excel_Location, Excel_Sheet_Name, Module_Name):
        data = {}
    print("Opening Excel")
    try:
        Excel_WorkBook = openpyxl.load_workbook(Excel_Location)
        print("Opened Excel")
    except Exception as error:
        print("Error in opening API Excel File")
        traceback.print_stack()

    try:
        for SheetName in Excel_WorkBook.worksheets:
            if SheetName.title == Excel_Sheet_Name:
                ActiveWorkSheet = Excel_WorkBook[SheetName.title]
                print("Title of Sheet = " + ActiveWorkSheet.title)
                RowLength = SheetName.max_row
                ColumnLength = SheetName.max_column
                for i in range(1, RowLength + 1):
                    RowContents = SheetName.cell(row=i, column=1)
                    print("Row " + str(RowContents.row) + " = " + str(RowContents.value))
                    if ((RowContents.value) == Module_Name):
                        print("Module Found : " + str(RowContents.value) + " in Row - " + str(
                            RowContents.row) + " of Worksheet - " + ActiveWorkSheet.title)
                    else:
                        continue

                    for j in range(1, ColumnLength + 1):
                        Response = dict()
                        ColumnContents = SheetName.cell(row=RowContents.row, column=j)
                        print(ColumnContents.value, end="" + "\n")

                        try:
                            API_Name = (SheetName["A" + str(RowContents.row)])
                            print(API_Name.value)
                            data['API'] = str(API_Name.value)
                            print(data['API'])
                        except Exception as error:
                            print(error)
                            traceback.print_stack()
                            Excel_WorkBook.close()
                            break

                        try:
                            HTTP_Method = (SheetName["B" + str(RowContents.row)])
                            print(HTTP_Method.value)
                            data['HTTPMethod'] = str(HTTP_Method.value)
                            print(data['HTTPMethod'])
                        except Exception as error:
                            print(error)
                            traceback.print_stack()
                            Excel_WorkBook.close()
                            break

                        try:
                            Request_Protocol = (SheetName["C" + str(RowContents.row)])
                            print(Request_Protocol.value)
                            data['Protocol'] = str(Request_Protocol.value)
                            print(data['Protocol'] + "\n")
                        except Exception as error:
                            print(error)
                            traceback.print_stack()
                            Excel_WorkBook.close()
                            break

                        try:
                            Request_BaseURL = (SheetName["D" + str(RowContents.row)])
                            print("BaseURL = " + str(Request_BaseURL.value) + "\n")
                        except Exception as error:
                            print(error)
                            traceback.print_stack()
                            Excel_WorkBook.close()
                            break

                        try:
                            Request_RelativeURL = (SheetName["E" + str(RowContents.row)])
                            print("RelativeURL = " + str(Request_RelativeURL.value) + "\n")
                        except Exception as error:
                            print(error)
                            traceback.print_stack()
                            Excel_WorkBook.close()
                            break

                        try:
                            data['URL'] = str(data['Protocol']) + "://" + str(Request_BaseURL.value) + str(Request_RelativeURL.value)
                            print(data['URL'])
                        except Exception as error:
                            print(error)
                            traceback.print_stack()
                            Excel_WorkBook.close()
                            break

                        try:
                            Body_Row = (SheetName["F" + str(RowContents.row)])
                            Body_Content = str(Body_Row.value)
                            print("Body = " + Body_Content + "\n")
                            data['Body'] = json.loads(Body_Content)
                            print(data['Body'])
                        except Exception as error:
                            print(error)
                            traceback.print_stack()
                            Excel_WorkBook.close()
                            break

                        try:
                            Header_Row = (SheetName["G" + str(RowContents.row)])
                            Header_Content = str(Header_Row.value)
                            print("Header = " + Header_Content + "\n")
                            data['Header'] = json.loads(Header_Content)
                            print(data['Header'])
                        except Exception as error:
                            print(error)
                            traceback.print_stack()
                            Excel_WorkBook.close()
                            break

                        try:
                            Cookie_Row = (SheetName["H" + str(RowContents.row)])
                            Cookie_Content = str(Cookie_Row.value)
                            print("Cookie = " + Cookie_Content + "\n")
                            data['Cookie'] = json.loads(Cookie_Content)
                            print(data['Cookie'])
                        except Exception as error:
                            print(error)
                            traceback.print_stack()
                            Excel_WorkBook.close()
                            break
                        break
                    break
                    print(data)
                    Excel_WorkBook.close()

    except:
        print("Error in reading contents")
    return data
"""