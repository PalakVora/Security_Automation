import re, json, time, traceback
import openpyxl
import requests
#from py._error import Error
import sys
sys.path.insert(1, 'PythonFiles/Utilities')
import Get_AttackAPI
import Get_AttackPosition
#import Read_Excel 

attack_payload_sheetname = "SQL"
excel_location =  "D:/Programming/Application Security/coe-application-security/DataFiles/API.xlsx"
payload_excel_location = "D:/Programming/Application Security/coe-application-security/DataFiles/Payloads.xlsx"
excel_sheetname = "ITSM"  
module_name = "TestAPI"
#Excel_WorkBook = openpyxl.load_workbook(excel_location)
def perform_attack(attack_area,http_method,any_parameter,payload_rowlength,payload_sheetname,url,body,header,set_cookie):
    result = {}
    if(any_parameter):
        print("Parameter found in = " + attack_area)
        for i in range(2, payload_rowlength + 1):
            payload_rowcontents = payload_sheetname.cell(row=i, column=1)
            print("Row " + str(payload_rowcontents.row - 1) + " = " + str(payload_rowcontents.value), end="" + "\n")
            for key in any_parameter:
                if(attack_area=='URL'):
                    print("Area is = " + attack_area)
                    AttackURL = url.replace(key, str(payload_rowcontents.value))
                    AttackURL = AttackURL.replace("$", "")
                    print("Attack url : " + str(AttackURL))
                    AttackBody = str(body).replace("$", "")
                    print("Body : " + str(AttackBody))
                    AttackHeader = str(header).replace("$", "")
                    print("Header : " + str(AttackHeader))
                    AttackCookie = str(set_cookie).replace("$", "")
                    print("Cookie : " + AttackCookie)
                elif(attack_area == 'Body'):
                    print("Area is = " + attack_area)
                    AttackBody = body.replace(key, str(payload_rowcontents.value))
                    print("Original BOdy ===============" + AttackBody)
                    print("for2", i+1)
                    AttackURL = url
                    print("URL : " + str(AttackURL))
                    AttackBody = str(AttackBody).replace("$", "")
                    print("AttackBody : " + str(AttackBody))
                    AttackHeader = str(header)
                    print("Header : " + str(AttackHeader))
                    AttackCookie = str(set_cookie)
                    print("Cookie : " +  AttackCookie)
                elif(attack_area == 'Header'):
                    print("Area is = " + attack_area)
                    AttackHeader = header.replace(key, str(payload_rowcontents.value))
                    AttackURL = AttackURL.replace("$", "")
                    print(AttackURL)
                    AttackBody = str(body).replace("$", "")
                    print(AttackBody)
                    AttackHeader = str(AttackHeader).replace("$", "")
                    print(AttackHeader)
                    AttackCookie = str(set_cookie).replace("$", "")
                    print(AttackCookie)
                elif(attack_area == 'Cookie'):
                    print("Area is = " + attack_area)
                    AttackCookie = set_cookie.replace(key, str(payload_rowcontents.value))
                    AttackURL = AttackURL.replace("$", "")
                    print(AttackURL)
                    AttackBody = str(body).replace("$", "")
                    print(AttackBody)
                    AttackHeader = str(header).replace("$", "")
                    print(AttackHeader)
                    AttackCookie = str(AttackCookie).replace("$", "")
                    print(AttackCookie)                
                #try:
                if (http_method == 'GET'):
                    print("Method found in attack = " + http_method)
                    print("Attack URl")
                    print(AttackURL)
                    print(type(AttackHeader))
                    response1 = requests.get(AttackURL, data=AttackBody, headers=AttackHeader, cookies=AttackCookie)
                    print(response1)
                    if response1:
                        print(response1.status_code)
                    else:
                        print(response1.status_code)
                elif(http_method == 'POST'):
                    print("Method found in attack = " + http_method)
                    response1 = requests.post(AttackURL, data=AttackBody, headers=AttackHeader,cookies=AttackCookie)
                    print("Got ============ response")
                elif(http_method == 'PUT'):
                    response1 = requests.put(AttackURL, data=AttackBody, headers=AttackHeader,cookies=AttackCookie)
                #elif(http_method == 'DELETE'):
                 #   response = requests.delete(AttackURL, data=AttackBody, headers=AttackHeader,cookies=AttackCookie)
                
    
    else:
        print("No Parameter choosen in the API")
    return result


#========================== MAIN FUNCTION WHICH IS CALLED BY API.ROBOT ======================================

def input_validation(excel_location, excel_sheetname, module_name, payload_excel_location, attack_payload_sheetname):
    payload_rowlength = 0
    payload_activeworksheet = ""
    api_hit_result = []
    try:
        result = Get_AttackAPI.find_api(excel_location, excel_sheetname, module_name)
        print("Data from find_vulnerable_parameters ")
        print(result)
    except:
        print('Error in executing Read Excel Method')

    attack = Get_AttackPosition.find_vulnerable_parameters(result)
    attack_method = attack['HTTPMethod']
    print("MEthod" + attack_method)

    attack_url = attack['URL']
    url_parameter = attack['URL_Parameter']  

    body = attack['Body']
    body_parameter = attack['Body_Parameter']

    header = attack['Header']
    header_parameter = attack['Header_Parameter']

    set_cookie = attack['Cookie']
    cookie_parameter = attack['Cookie_Parameter']
    try:
        payload_excel_workbook_location = openpyxl.load_workbook(payload_excel_location)
        for payload_sheetname in payload_excel_workbook_location.worksheets:
            if payload_sheetname.title == attack_payload_sheetname:
                print("Worksheet Found = " + payload_sheetname.title)
                payload_activeworksheet = payload_excel_workbook_location[payload_sheetname.title]
                print("Title of Sheet = " + payload_activeworksheet.title)
                payload_rowlength = payload_sheetname.max_row
                print("Number of Payloads for " + str(payload_sheetname.title) + " = " + str(payload_rowlength - 1))
        if(attack_method == 'GET'):
            print('Executing GET Method')
            if(url_parameter):
                print("Parameter found in URL")
                api_hit_result = perform_attack('URL','GET',url_parameter,payload_rowlength,payload_activeworksheet,attack_url,body,header,set_cookie)
            elif(body_parameter):
                print("Parameter found in Body")
                api_hit_result = perform_attack('Body','GET',body_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)

            elif(header_parameter):
                print("Parameter found in POST Header")
                api_hit_result = perform_attack('Header','GET',header_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)

            elif(cookie_parameter):
                print("Parameter found in POST Cookie")
                api_hit_result = perform_attack('Cookie','GET',cookie_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)
                    
            else:
                print("No Parameter choosen in the API")
                
        elif(attack_method == 'POST'):
            print('Executing POST Method')

            if(url_parameter):
                print("Parameter found in URL")
                api_hit_result = perform_attack('URL','POST',url_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)
            
            elif(body_parameter):
                print("Parameter found in POST Body")
                api_hit_result = perform_attack('Body','POST',body_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)

            elif(header_parameter):
                print("Parameter found in POST Header")
                api_hit_result = perform_attack('Header','POST',header_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)

            elif(cookie_parameter):
                print("Parameter found in POST Cookie")
                api_hit_result = perform_attack('Cookie','POST',cookie_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)
                    
            else:
                print("No Parameter choosen in the API")

        elif(attack_method == 'PUT'):
            print('Executing PUT Method')

            if(url_parameter):
                print("Parameter found in PUT URL")
                api_hit_result = perform_attack('URL','PUT',url_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)


            elif(body_parameter):
                print("Parameter found in PUT Body")
                api_hit_result = perform_attack('Body','PUT',body_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)


            elif(header_parameter):
                print("Parameter found in PUT Header")
                api_hit_result = perform_attack('Header','PUT',header_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)

        
            elif(cookie_parameter):
                print("Parameter found in PUT Cookie")
                api_hit_result = perform_attack('Cookie','PUT',cookie_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)
        
            else:
                print("No Parameter choosen in the API")

        if(attack_method == 'DELETE'):
            print('Executing DELETE Method')

            if(url_parameter):
                print("Parameter found in DELETE URL")
                api_hit_result = perform_attack('URL','DELETE',url_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)


            elif(body_parameter):
                print("Parameter found in DELETE Body")
                api_hit_result = perform_attack('Body','DELETE',body_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)


            elif(header_parameter):
                print("Parameter found in DELETE Header")
                api_hit_result = perform_attack('Header','DELETE',header_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)

        
            elif(cookie_parameter):
                print("Parameter found in DELETE Cookie")
                api_hit_result = perform_attack('Cookie','DELETE',cookie_parameter,payload_rowlength,payload_sheetname,attack_url,body,header,set_cookie)
        
            else:
                print("No Parameter choosen in the API")

    except:
        print("Error in identifying the method")
    return api_hit_result

if __name__ == "__main__":
    getit = input_validation(excel_location, excel_sheetname, module_name, payload_excel_location, attack_payload_sheetname)
    print("WELL++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
    print(getit)
"""
    for key in URL_Parameter:
        AttackURL = URL.replace(key, '<script>alert(1)</script>')
        print(AttackURL)
        AttackURL = AttackURL.replace("$", "")
        AttackBody = str(Body).replace("$", "")
        AttackHeader = str(Header).replace("$", "")
        AttackCookie = str(Cookie).replace("$", "")
        print("Vulnerable URL = " + AttackURL)
        try:
            response = requests.post(str(AttackURL), data=AttackBody, headers=AttackHeader, cookies=AttackCookie, timeout=100)
            print(response)
            StatusCode = str(response.status_code)
            result['StatusCode'] = str(StatusCode)
        except:
            print("Error in executing: " + str(AttackURL))
            result['StatusCode'] = '500'


    print(result)
    return result



        for key in Parametrized_URL:
            newURL = URL.replace(key, '<script>alert(1)</script>')
            newURL = newURL.replace("$", "")
            print("Vulnerable URL = " + newURL)
            break



          #except:
                 #   print(traceback)
                  #  print("Error in executing: " + str(AttackURL))
                   # StatusCode = '500'
                   # result.append(StatusCode)
                    #print(result)
                   # time.sleep(10)
               # print(result)



               StatusCode = str(response.status_code)
                Response_Body = str(response.text)
                print("Response Status Code : " + str(StatusCode) + "\n")
                print("Response Body : " + str(Response_Body) + "\n")
                result.append(StatusCode)
                print(result)
                time.sleep(10)



    '''   '''             
"""