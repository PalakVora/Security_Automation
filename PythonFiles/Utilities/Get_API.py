import openpyxl,traceback,json


class ReadApi:
   
    Excel_Sheet_Name=""
    module_name=""
    Excel_Location=""
    def __init__(self):
        self.rowcontents = ""
        self.sheetname = ""
        self.excel_workbook=""
        self.request_base_url=""
        self.body_row=""
        self.cookie_row=""
        self.header_row=""
        self.request_protocol=""
        self.request_relative_url=""
        self.http_method=""
        self.api_name=""

    def _dir__(self):
        return (self.request_base_url, str(self.body_row.value),str(self.cookie_row.value),str(self.header_row.value),str(self.request_protocol.value),self.request_relative_url,str(self.http_method.value))

    def readexcel(self):
        data = {}
        print("Opening Excel")
        try:
            self.excel_workbook = openpyxl.load_workbook(self.Excel_Location)
            print("Opened Excel")
        except Exception as error:
            print("Error in opening API Excel File")
            traceback.print_stack()
        #return Excel_WorkBook

        #========================== GET SHEET FROM EXCEL======================================

    def find_sheet_name(self):
        print("inside sheet name")
        for sheet in self.excel_workbook.worksheets:
            if sheet.title == self.Excel_Sheet_Name:
                ActiveWorkSheet = self.excel_workbook[sheet.title]
                print("Title of Sheet = " + ActiveWorkSheet.title)
                print("Sheetname =" + sheet.title)
                self.sheetname=sheet
                break
               
            else:
                continue
    
#========================== FIND API NAME FROM EXCEL SHEET ======================================

    def find_module_name(self):
        rowlength = self.sheetname.max_row
        for i in range(1, rowlength + 1):
            self.rowcontents = self.sheetname.cell(row=i, column=1)
            print("Row " + str(self.rowcontents.row) + " = " + str(self.rowcontents.value))
            if ((self.rowcontents.value) == self.module_name):
                print("Module Found : " + str(self.rowcontents.value) + " in Row - " + str(self.rowcontents.row) + " of Worksheet - " + self.sheetname.title)
                return self.rowcontents
            else:
                continue
            try:
                print(str((self.sheetname["A" + str(self.rowcontents.row)]).value) + "\n That API name section")
            except:
                print("Error in getting API Name")
                traceback.print_stack()
            break


    def read_api_name(self):
        try:
            self.api_name = (self.sheetname["A" + str(self.rowcontents.row)])
        except:
            print("Error in api name")
                
 # ------------------------------------ Read Base URL ------------------------             
    def read_base_url(self):
        print(str(self.rowcontents.row))
        try:
            self.request_base_url = (self.sheetname["D" + str(self.rowcontents.row)])
        except:
            print("Error in reading Request Base URL from Excel File")
            self.excel_workbook.close() 

# ------------------------- Read BODY---------------------
    def read_body(self):
        try:
            self.body_row = (self.sheetname["F" + str(self.rowcontents.row)])
            self.body_row = str(self.body_row.value)
            #return str(body_row.value)
        except:
            print("Error in reading Request Body from Excel File")
            self.excel_workbook.close()

# ----------------------- Get Cookie---------------------------------
    def read_cookie(self):
        try:
            self.cookie_row = (self.sheetname["H" + str(self.rowcontents.row)])
            self.cookie_row = str(self.cookie_row.value)
            #return str(cookie_row.value)
        except:
            print("Error in reading Request Cookie from Excel File")
            self.excel_workbook.close()

# ------------------------------ Get Header -------------------------
    def read_header(self):
        try:
            self.header_row = (self.sheetname["G" + str(self.rowcontents.row)])
            self.header_row = str(self.header_row.value)
            #return str(header_row.value)
        except:
            print("Error in reading Request Header from Excel File")
            self.excel_workbook.close()      


# ------------------------------ Get Protocol -------------------------
    def read_protocol(self):
        try:
            self.request_protocol = (self.sheetname["C" + str(self.rowcontents.row)])
            #return str(request_protocol.value)
        except:
            print("Error in reading Request Protocol from Excel File")
            self.excel_workbook.close()     

# ------------------------------ Get Relative URL -------------------------
    def read_relative_url(self):   
        try:
             self.request_relative_url = (self.sheetname["E" + str(self.rowcontents.row)])
             print("RelativeURL = " + str(self.request_relative_url.value) + "\n")
             #return request_relative_url
        except:
            print("Error in reading Request Relative URL from Excel File")
            self.excel_workbook.close()

# ------------------------------ Get Verb -------------------------
    def read_method(self):
            try:
                self.http_method = (self.sheetname["B" + str(self.rowcontents.row)])
                print("ROw" + str(self.rowcontents.row))
                print(self.http_method.value)
                #return str(http_method.value)
            except Exception as error:
                print("Error in reading HTTP Method from Excel File")
                print(error)
                self.excel_workbook.close()
                #return 0
                        

